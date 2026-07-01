"""
ETL para cargar Indicadores Históricos al sistema.

Uso local (sin BD):
  python scripts/cargar_indicadores.py --dry-run
  python scripts/cargar_indicadores.py --export-csv ./salida

Uso en contenedor backend:
  docker cp scripts/cargar_indicadores.py academia-stats-backend-1:/app/
  docker cp "Indicadores historicos 07-02-25 (2).xlsx" academia-stats-backend-1:/app/
  docker exec academia-stats-backend-1 python3 /app/cargar_indicadores.py
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_EXCEL = Path.home() / "Downloads/Indicadores historicos 07-02-25 (2).xlsx"
if not DEFAULT_EXCEL.exists():
    DEFAULT_EXCEL = Path("/app/Indicadores historicos 07-02-25 (2).xlsx")

_SUBSISTEMA_ID = 5  # Universidades Politécnicas


def _sid() -> int:
    return _SUBSISTEMA_ID

PE_ALIASES = {"LADM": "LAD", "ITIID": "ITID", "LAGP": "LAGP"}
SKIP_PE = {"MATRICULA", "TOTAL", "ACEPTADOS"}
PE_PATTERN = re.compile(r"^[A-Z]{2,6}$")

MATRICULA_SHEET = "Matricula por cuatrimestre"
NUEVO_INGRESO_SHEET = "Matricula H. Nuevo Ingreso"
GENERO_SHEET = "Matricula por Género"
REP_DES_SHEET = "REP Y DES"
APROVECHAMIENTO_SHEET = "APROVECHAMIENTO"
TITULACION_SHEET = "Titulados Histórico act"

APROVECHAMIENTO_YEARS = list(range(2015, 2025))

EVAL_DOC_SHEETS: dict[str, tuple[str, int]] = {
    "EVAL DOC CLIENTE Y RPE S-D 20": ("2020-2021", 1),
    "EVAL DOC CLIENTE Y RPE E-A 21": ("2020-2021", 2),
    "EVAL DOC CLIENTE Y RPE M-A 21": ("2020-2021", 3),
    "EVAL DOC CLIENTE Y RPE E-A 22": ("2021-2022", 2),
    "EVAL DOC CLIENTE Y RPE M-A 22": ("2021-2022", 3),
    "EVAL DOC CLIENTE Y RPE  M-A 23": ("2022-2023", 3),
    "EVAL DOC CLIENTE Y RPE  E-A 24": ("2023-2024", 2),
    "EVAL DOC CLIENTE Y RPE  M-A 24": ("2023-2024", 3),
}

MatriculaKey = tuple[str, int, str]


# ── Helpers ───────────────────────────────────────────────────────────────────

def _normalize_id(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", "_", s.strip().lower())


def _normalize_pe(pe: str) -> str:
    pe = str(pe).strip().upper()
    return PE_ALIASES.get(pe, pe)


def _parse_periodo(label: str) -> tuple[str, int] | tuple[None, None]:
    label = str(label).strip()
    m = re.search(r"\d{4}", label)
    if not m:
        return None, None
    year = int(m.group())
    lo = label.lower().replace(" ", "")
    if any(x in lo for x in ("sep", "s-d", "sd", "s_d")):
        return f"{year}-{year + 1}", 1
    if any(x in lo for x in ("ene", "e-a", "ea", "e_a")):
        return f"{year - 1}-{year}", 2
    if any(x in lo for x in ("may", "m-a", "ma", "m_a")):
        return f"{year - 1}-{year}", 3
    return None, None


def _safe_int(v: Any) -> int:
    try:
        if v is None:
            return 0
        return int(float(str(v)))
    except (TypeError, ValueError):
        return 0


def _safe_float(v: Any) -> float | None:
    try:
        if v is None:
            return None
        f = float(str(v))
        if f <= 0:
            return None
        return round(f, 4)
    except (TypeError, ValueError):
        return None


def _expand_merged(ws: Any) -> None:
    for rng in list(ws.merged_cells.ranges):
        val = ws.cell(rng.min_row, rng.min_col).value
        ws.unmerge_cells(str(rng))
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                ws.cell(r, c).value = val


def _row_vals(ws: Any, row: int) -> dict[int, Any]:
    return {
        (cell.column - 1): cell.value
        for cell in ws[row]
        if cell.value is not None
    }


def _find_header_row(ws: Any, marker: str = "PE", max_row: int = 15) -> int | None:
    for r in ws.iter_rows(max_row=max_row):
        vals = [c.value for c in r if c.value is not None]
        if any(str(v).strip().upper() == marker for v in vals):
            return r[0].row
    return None


def _matricula_key(ciclo: str, cuatrimestre: int, pe: str) -> MatriculaKey:
    return (ciclo, cuatrimestre, _normalize_pe(pe))


def _new_matricula_row(ciclo: str, cuatrimestre: int, pe: str, total: int = 0) -> dict:
    return {
        "subsistema_id": _sid(),
        "ciclo_escolar": ciclo,
        "cuatrimestre": cuatrimestre,
        "programa_educativo": _normalize_pe(pe),
        "total": total,
        "nuevo_ingreso": 0,
        "bajas_reprobacion": 0,
        "bajas_desercion": 0,
        "hombres": 0,
        "mujeres": 0,
        "poblacion_edad_escolar": None,
        "egresados_nms": None,
    }


def _distribute_proportional(
    records: dict[MatriculaKey, dict],
    ciclo: str,
    cuatrimestre: int,
    total_value: int,
    field: str,
) -> None:
    if total_value <= 0:
        return
    period_rows = [
        (k, r) for k, r in records.items()
        if k[0] == ciclo and k[1] == cuatrimestre and r["total"] > 0
    ]
    if not period_rows:
        return
    weight_total = sum(r["total"] for _, r in period_rows)
    if weight_total <= 0:
        return

    shares: list[tuple[MatriculaKey, int]] = []
    allocated = 0
    for key, row in period_rows:
        share = round(total_value * row["total"] / weight_total)
        shares.append((key, share))
        allocated += share

    diff = total_value - allocated
    if diff != 0:
        largest = max(shares, key=lambda x: records[x[0]]["total"])
        shares = [(k, v + (diff if k == largest[0] else 0)) for k, v in shares]

    for key, share in shares:
        records[key][field] = share


def _parse_sheet_period_columns(ws: Any, header_row: int) -> dict[int, tuple[str, int]]:
    header = _row_vals(ws, header_row)
    col_map: dict[int, tuple[str, int]] = {}
    for col_idx, val in header.items():
        if val is None or str(val).strip().upper() == "PE":
            continue
        ciclo, cuatrimestre = _parse_periodo(str(val))
        if ciclo and cuatrimestre:
            col_map[col_idx] = (ciclo, cuatrimestre)
    return col_map


# ── 1. MATRÍCULA ─────────────────────────────────────────────────────────────

def extract_matricula_base(wb: openpyxl.Workbook) -> dict[MatriculaKey, dict]:
    if MATRICULA_SHEET not in wb.sheetnames:
        print(f"  [WARN] Hoja '{MATRICULA_SHEET}' no encontrada")
        return {}

    ws = wb[MATRICULA_SHEET]
    _expand_merged(ws)
    header_row = _find_header_row(ws)
    if header_row is None:
        print(f"  [WARN] {MATRICULA_SHEET}: sin fila de encabezados")
        return {}

    col_map = _parse_sheet_period_columns(ws, header_row)
    records: dict[MatriculaKey, dict] = {}

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        pe = row[1] if len(row) > 1 else None
        if pe is None:
            continue
        pe_str = str(pe).strip().upper()
        if pe_str in SKIP_PE or not PE_PATTERN.match(_normalize_pe(pe_str)):
            continue

        for col_idx, (ciclo, cuatrimestre) in col_map.items():
            total = _safe_int(row[col_idx] if col_idx < len(row) else None)
            if total <= 0:
                continue
            key = _matricula_key(ciclo, cuatrimestre, pe_str)
            if key in records:
                records[key]["total"] += total
            else:
                records[key] = _new_matricula_row(ciclo, cuatrimestre, pe_str, total)

    return records


def _apply_nuevo_ingreso(wb: openpyxl.Workbook, records: dict[MatriculaKey, dict]) -> None:
    if NUEVO_INGRESO_SHEET not in wb.sheetnames:
        return
    ws = wb[NUEVO_INGRESO_SHEET]
    _expand_merged(ws)

    header_row = _find_header_row(ws, marker="PE", max_row=10)
    if header_row is None:
        return

    col_map: dict[int, tuple[str, int]] = {}
    for col_idx, val in _row_vals(ws, header_row).items():
        if str(val).strip().upper() == "PE":
            continue
        lo = str(val).lower().replace(" ", "")
        if not any(x in lo for x in ("s-d", "sd", "sep")):
            continue
        ciclo, cuatrimestre = _parse_periodo(str(val))
        if ciclo and cuatrimestre == 1:
            col_map[col_idx] = (ciclo, cuatrimestre)

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        pe = row[1] if len(row) > 1 else None
        if pe is None:
            continue
        pe_str = str(pe).strip().upper()
        if pe_str in SKIP_PE or not PE_PATTERN.match(_normalize_pe(pe_str)):
            continue

        for col_idx, (ciclo, cuatrimestre) in col_map.items():
            ni = _safe_int(row[col_idx] if col_idx < len(row) else None)
            if ni <= 0:
                continue
            key = _matricula_key(ciclo, cuatrimestre, pe_str)
            if key not in records:
                records[key] = _new_matricula_row(ciclo, cuatrimestre, pe_str)
            records[key]["nuevo_ingreso"] = ni


def _apply_genero(wb: openpyxl.Workbook, records: dict[MatriculaKey, dict]) -> None:
    if GENERO_SHEET not in wb.sheetnames:
        return
    ws = wb[GENERO_SHEET]
    _expand_merged(ws)

    header_row = None
    for r in ws.iter_rows(max_row=15):
        vals = [str(c.value).lower() for c in r if c.value is not None]
        if any("sep" in v or "ene" in v or "may" in v for v in vals):
            header_row = r[0].row
            break
    if header_row is None:
        return

    col_map = _parse_sheet_period_columns(ws, header_row)
    mujeres_row = hombres_row = None
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 5, values_only=True):
        label = str(row[1] if len(row) > 1 else "").strip().lower()
        if "mujer" in label:
            mujeres_row = row
        elif "hombre" in label:
            hombres_row = row

    if not mujeres_row or not hombres_row:
        return

    for col_idx, (ciclo, cuatrimestre) in col_map.items():
        mujeres = _safe_int(mujeres_row[col_idx] if col_idx < len(mujeres_row) else None)
        hombres = _safe_int(hombres_row[col_idx] if col_idx < len(hombres_row) else None)
        if mujeres > 0:
            _distribute_proportional(records, ciclo, cuatrimestre, mujeres, "mujeres")
        if hombres > 0:
            _distribute_proportional(records, ciclo, cuatrimestre, hombres, "hombres")


def _apply_rep_des(wb: openpyxl.Workbook, records: dict[MatriculaKey, dict]) -> None:
    if REP_DES_SHEET not in wb.sheetnames:
        return
    ws = wb[REP_DES_SHEET]
    _expand_merged(ws)

    header_row = None
    period_cols: dict[int, tuple[str, int]] = {}
    for r in ws.iter_rows(max_row=10):
        vals = [c.value for c in r if c.value is not None]
        if len(vals) >= 3 and any(_parse_periodo(str(v))[0] for v in vals):
            header_row = r[0].row
            for col_idx, val in _row_vals(ws, header_row).items():
                ciclo, cuat = _parse_periodo(str(val))
                if ciclo and cuat:
                    period_cols[col_idx] = (ciclo, cuat)
            break
    if header_row is None:
        return

    metrics: dict[str, dict[int, int]] = {}
    for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 6, values_only=True):
        label = str(row[0] if row else "").strip().lower()
        if "deserci" in label:
            metrics["bajas_desercion"] = {
                c: _safe_int(row[c] if c < len(row) else None) for c in period_cols
            }
        elif "reprobaci" in label and "%" not in label:
            metrics["bajas_reprobacion"] = {
                c: _safe_int(row[c] if c < len(row) else None) for c in period_cols
            }
        elif "egresado" in label:
            metrics["egresados_nms"] = {
                c: _safe_int(row[c] if c < len(row) else None) for c in period_cols
            }

    for field, col_values in metrics.items():
        for col_idx, value in col_values.items():
            if value <= 0:
                continue
            ciclo, cuatrimestre = period_cols[col_idx]
            if field == "egresados_nms":
                period_rows = [
                    (k, r) for k, r in records.items()
                    if k[0] == ciclo and k[1] == cuatrimestre and r["total"] > 0
                ]
                if not period_rows:
                    continue
                weight_total = sum(r["total"] for _, r in period_rows)
                shares = []
                allocated = 0
                for key, row in period_rows:
                    share = round(value * row["total"] / weight_total)
                    shares.append((key, share))
                    allocated += share
                diff = value - allocated
                if diff and shares:
                    largest = max(shares, key=lambda x: records[x[0]]["total"])
                    shares = [(k, v + (diff if k == largest[0] else 0)) for k, v in shares]
                for key, share in shares:
                    records[key]["egresados_nms"] = share
            else:
                _distribute_proportional(records, ciclo, cuatrimestre, value, field)


def extract_matricula(wb: openpyxl.Workbook) -> list[dict]:
    records = extract_matricula_base(wb)
    _apply_nuevo_ingreso(wb, records)
    _apply_genero(wb, records)
    _apply_rep_des(wb, records)
    return list(records.values())


# ── 2. EVALUACIÓN ACADÉMICA ───────────────────────────────────────────────────

def extract_evaluacion_academica(wb: openpyxl.Workbook) -> list[dict]:
    if APROVECHAMIENTO_SHEET not in wb.sheetnames:
        return []
    ws = wb[APROVECHAMIENTO_SHEET]
    _expand_merged(ws)

    rows: list[dict] = []
    for r in ws.iter_rows(min_row=6, max_row=20):
        pe_val = r[1].value if len(r) > 1 else None
        if not pe_val or not isinstance(pe_val, str):
            continue
        pe = _normalize_pe(pe_val.strip())
        if not PE_PATTERN.match(pe) or pe in SKIP_PE:
            continue

        row_data = list(ws.iter_rows(min_row=r[0].row, max_row=r[0].row, values_only=True))[0]
        for y_offset, year in enumerate(APROVECHAMIENTO_YEARS):
            base_col = 3 + y_offset * 3
            cuatrimestre_map = [
                (base_col, year - 1, year, 2),
                (base_col + 1, year - 1, year, 3),
                (base_col + 2, year, year + 1, 1),
            ]
            for col_idx, cy1, cy2, cuatrimestre in cuatrimestre_map:
                if col_idx >= len(row_data):
                    continue
                prom = _safe_float(row_data[col_idx])
                if prom is None:
                    continue
                rows.append({
                    "subsistema_id": _sid(),
                    "ciclo_escolar": f"{cy1}-{cy2}",
                    "cuatrimestre": cuatrimestre,
                    "programa_educativo": pe,
                    "promedio_pe": prom,
                    "num_pe": 1,
                })
    return rows


# ── 3. EVALUACIÓN DOCENTE ─────────────────────────────────────────────────────

def _find_eval_doc_columns(header_row: tuple) -> dict[str, int]:
    cols: dict[str, int] = {}
    for idx, val in enumerate(header_row):
        if val is None:
            continue
        norm = str(val).strip().lower().replace("\n", " ")
        if "profesor" in norm and "profesor" not in cols:
            cols["profesor"] = idx
        elif norm in ("estudiante", "estudiantes") or "estudiante" in norm:
            cols["estudiante"] = idx
        elif norm == "rpe":
            cols["rpe"] = idx
    return cols


def _is_docente_name(val: Any) -> bool:
    if not isinstance(val, str):
        return False
    name = val.strip()
    if len(name) < 5:
        return False
    if any(x in name.upper() for x in ("DOCENTES", "EVALUADOS", "TOTAL", "MAL", "REGULAR")):
        return False
    return bool(re.search(r"[A-Za-zÁÉÍÓÚáéíóúñÑ]{3,}", name))


def _extract_eval_doc_sheet(ws: Any, ciclo: str) -> list[dict]:
    rows: list[dict] = []
    header_idx = None
    col_map: dict[str, int] = {}

    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=True), start=1):
        cmap = _find_eval_doc_columns(row)
        if "profesor" in cmap and ("estudiante" in cmap or "rpe" in cmap):
            header_idx = i
            col_map = cmap
            break
    if header_idx is None:
        return rows

    p_col = col_map["profesor"]
    e_col = col_map.get("estudiante")
    r_col = col_map.get("rpe")

    for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
        nombre = row[p_col] if p_col < len(row) else None
        if not _is_docente_name(nombre):
            continue
        nombre = str(nombre).strip()
        docente_id = _normalize_id(nombre)

        if e_col is not None and e_col < len(row):
            est = _safe_float(row[e_col])
            if est is not None:
                rows.append({
                    "subsistema_id": _sid(),
                    "ciclo_escolar": ciclo,
                    "docente_id": docente_id,
                    "docente_nombre": nombre,
                    "programa_educativo": "GENERAL",
                    "evaluador_tipo": "alumno",
                    "puntaje": est,
                })
        if r_col is not None and r_col < len(row):
            rpe = _safe_float(row[r_col])
            if rpe is not None:
                rows.append({
                    "subsistema_id": _sid(),
                    "ciclo_escolar": ciclo,
                    "docente_id": docente_id,
                    "docente_nombre": nombre,
                    "programa_educativo": "GENERAL",
                    "evaluador_tipo": "directivo",
                    "puntaje": rpe,
                })
    return rows


def _extract_eval_doc_dual_2020(ws: Any) -> list[dict]:
    """Hoja E-A / M-A 2020 con dos bloques de columnas."""
    rows: list[dict] = []
    header = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]

    blocks: list[tuple[str, dict[str, int]]] = []
    prof_indices = [i for i, v in enumerate(header) if v and "profesor" in str(v).lower()]

    for pi, p_idx in enumerate(prof_indices):
        block: dict[str, int] = {"profesor": p_idx}
        end = prof_indices[pi + 1] if pi + 1 < len(prof_indices) else len(header)
        for j in range(p_idx, end):
            val = header[j]
            if val is None:
                continue
            norm = str(val).lower()
            if "e-a" in norm or "e-a" in norm.replace(" ", ""):
                block["ciclo"] = "2019-2020"
            elif "m-a" in norm or "m-a" in norm.replace(" ", ""):
                block["ciclo"] = "2019-2020"
            if j > p_idx and _safe_float(val) and "ciclo" not in block:
                pass
            if norm.strip() in ("estudiante",) or "estudiante" in norm:
                block["estudiante"] = j
            elif norm.strip() == "rpe":
                block["rpe"] = j
        if "estudiante" in block or "rpe" in block:
            if "ciclo" not in block:
                block["ciclo"] = "2019-2020" if pi == 0 else "2019-2020"
            blocks.append((block.get("ciclo", "2019-2020"), block))

    if not blocks:
        return _extract_eval_doc_sheet(ws, "2019-2020")

    for ciclo, block in blocks:
        for row in ws.iter_rows(min_row=4, values_only=True):
            nombre = row[block["profesor"]] if block["profesor"] < len(row) else None
            if not _is_docente_name(nombre):
                continue
            nombre = str(nombre).strip()
            docente_id = _normalize_id(nombre)
            if "estudiante" in block:
                est = _safe_float(row[block["estudiante"]] if block["estudiante"] < len(row) else None)
                if est is not None:
                    rows.append({
                        "subsistema_id": _sid(),
                        "ciclo_escolar": ciclo,
                        "docente_id": docente_id,
                        "docente_nombre": nombre,
                        "programa_educativo": "GENERAL",
                        "evaluador_tipo": "alumno",
                        "puntaje": est,
                    })
            if "rpe" in block:
                rpe = _safe_float(row[block["rpe"]] if block["rpe"] < len(row) else None)
                if rpe is not None:
                    rows.append({
                        "subsistema_id": _sid(),
                        "ciclo_escolar": ciclo,
                        "docente_id": docente_id,
                        "docente_nombre": nombre,
                        "programa_educativo": "GENERAL",
                        "evaluador_tipo": "directivo",
                        "puntaje": rpe,
                    })
    return rows


def extract_evaluacion_docente(wb: openpyxl.Workbook) -> list[dict]:
    rows: list[dict] = []
    dual_sheet = "EVAL DOC CLIENTE E-A M-A 2020"
    if dual_sheet in wb.sheetnames:
        rows.extend(_extract_eval_doc_dual_2020(wb[dual_sheet]))

    for sheet_name, (ciclo, _cuat) in EVAL_DOC_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        rows.extend(_extract_eval_doc_sheet(wb[sheet_name], ciclo))

    return _dedup_rows(rows, ["subsistema_id", "ciclo_escolar", "docente_id", "evaluador_tipo"])


# ── 4. TITULACIÓN ─────────────────────────────────────────────────────────────

def extract_titulacion(wb: openpyxl.Workbook) -> list[dict]:
    if TITULACION_SHEET not in wb.sheetnames:
        return []
    ws = wb[TITULACION_SHEET]
    rows: list[dict] = []
    current_carrera: str | None = None

    for row in ws.iter_rows(min_row=15, values_only=True):
        if len(row) < 15:
            continue
        carrera_val = row[1]
        if carrera_val and isinstance(carrera_val, str) and len(carrera_val.strip()) > 5:
            current_carrera = carrera_val.strip()
        if current_carrera is None:
            continue
        ingreso_val = row[2]
        if not isinstance(ingreso_val, datetime):
            continue

        matricula_gen = _safe_int(row[6])
        egresados_total = _safe_int(row[11])
        titulados_total = _safe_int(row[14])
        if matricula_gen == 0:
            continue

        rows.append({
            "subsistema_id": _sid(),
            "generacion": f"{ingreso_val.year}-{ingreso_val.strftime('%m')}",
            "programa_educativo": current_carrera,
            "matricula_generacional": matricula_gen,
            "concluyeron_estudios": egresados_total,
            "egresados": egresados_total,
            "titulados": titulados_total,
            "ingresados_ns": None,
        })
    return rows


# ── Utilidades ────────────────────────────────────────────────────────────────

def _dedup_rows(rows: list[dict], keys: list[str]) -> list[dict]:
    seen: dict[tuple, dict] = {}
    for row in rows:
        k = tuple(row[key] for key in keys)
        seen[k] = row
    return list(seen.values())


def _validation_summary(matricula_rows: list[dict]) -> None:
    totals_by_period: dict[tuple[str, int], int] = defaultdict(int)
    ni_by_period: dict[tuple[str, int], int] = defaultdict(int)
    for r in matricula_rows:
        key = (r["ciclo_escolar"], r["cuatrimestre"])
        totals_by_period[key] += r["total"]
        ni_by_period[key] += r["nuevo_ingreso"]

    checks = [
        (("2022-2023", 1), 2113, "Matrícula S-D 2022"),
        (("2023-2024", 1), 2128, "Matrícula S-D 2023"),
    ]
    print("\n  Validación de totales institucionales:")
    for period, expected, label in checks:
        actual = totals_by_period.get(period, 0)
        ok = "OK" if actual == expected else f"DIFF ({actual} vs {expected})"
        print(f"    {label}: {ok}")

    sd_2024_ni = sum(
        r["nuevo_ingreso"] for r in matricula_rows
        if r["ciclo_escolar"] == "2024-2025" and r["cuatrimestre"] == 1
    )
    print(f"    Nuevo ingreso S-D 2024 (suma PE): {sd_2024_ni} (esperado ~815)")


MATRICULA_CSV_COLS = [
    "ciclo_escolar", "cuatrimestre", "programa_educativo", "total", "nuevo_ingreso",
    "bajas_reprobacion", "bajas_desercion", "hombres", "mujeres",
    "poblacion_edad_escolar", "egresados_nms",
]
EVAL_ACAD_CSV_COLS = [
    "ciclo_escolar", "cuatrimestre", "programa_educativo", "promedio_pe", "num_pe",
]
TITULACION_CSV_COLS = [
    "generacion", "programa_educativo", "matricula_generacional",
    "concluyeron_estudios", "egresados", "titulados", "ingresados_ns",
]
EVAL_DOC_CSV_COLS = [
    "ciclo_escolar", "docente_id", "docente_nombre", "programa_educativo",
    "evaluador_tipo", "puntaje",
]


def export_csv(
    out_dir: Path,
    matricula_rows: list[dict],
    eval_acad_rows: list[dict],
    eval_doc_rows: list[dict],
    tit_rows: list[dict],
) -> None:
    out_dir.mkdir(parents=True, exist_ok=True)
    datasets = [
        ("carga_matricula.csv", matricula_rows, MATRICULA_CSV_COLS),
        ("carga_evaluacion_academica.csv", eval_acad_rows, EVAL_ACAD_CSV_COLS),
        ("carga_titulacion.csv", tit_rows, TITULACION_CSV_COLS),
        ("carga_evaluacion_docente.csv", eval_doc_rows, EVAL_DOC_CSV_COLS),
    ]
    for filename, rows, cols in datasets:
        path = out_dir / filename
        with path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=cols, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                out: dict[str, Any] = {}
                for c in cols:
                    val = row.get(c)
                    if val is None:
                        out[c] = ""
                    elif c in ("poblacion_edad_escolar", "egresados_nms", "ingresados_ns") and val == "":
                        out[c] = ""
                    else:
                        out[c] = val
                writer.writerow(out)
        print(f"  ✓ {path} ({len(rows)} filas)")


def load_to_database(
    matricula_rows: list[dict],
    eval_acad_rows: list[dict],
    eval_doc_rows: list[dict],
    tit_rows: list[dict],
) -> None:
    from sqlalchemy import create_engine, text
    from sqlalchemy.dialects.postgresql import insert as pg_insert
    from sqlalchemy.orm import Session

    from app.core.config import settings
    from app.models.evaluacion import EvaluacionAcademica, EvaluacionDocente
    from app.models.matricula import Matricula
    from app.models.titulacion import Titulacion

    engine = create_engine(settings.database_url_str, pool_pre_ping=True, future=True)

    def upsert_matricula(session: Session, rows: list[dict]) -> int:
        if not rows:
            return 0
        rows = _dedup_rows(rows, ["subsistema_id", "ciclo_escolar", "cuatrimestre", "programa_educativo"])
        stmt = pg_insert(Matricula.__table__).values(rows)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_matricula_periodo",
            set_={c: stmt.excluded[c] for c in [
                "total", "nuevo_ingreso", "bajas_reprobacion", "bajas_desercion",
                "hombres", "mujeres", "egresados_nms",
            ]},
        )
        return session.execute(stmt).rowcount or len(rows)

    def insert_eval_academica(session: Session, rows: list[dict]) -> int:
        if not rows:
            return 0
        session.execute(
            text("DELETE FROM evaluacion_academica WHERE subsistema_id = :sid"),
            {"sid": _sid()},
        )
        session.execute(pg_insert(EvaluacionAcademica.__table__).values(rows))
        return len(rows)

    def insert_eval_docente(session: Session, rows: list[dict]) -> int:
        if not rows:
            return 0
        session.execute(
            text("DELETE FROM evaluacion_docente WHERE subsistema_id = :sid"),
            {"sid": _sid()},
        )
        session.execute(pg_insert(EvaluacionDocente.__table__).values(rows))
        return len(rows)

    def upsert_titulacion(session: Session, rows: list[dict]) -> int:
        if not rows:
            return 0
        rows = _dedup_rows(rows, ["subsistema_id", "generacion", "programa_educativo"])
        stmt = pg_insert(Titulacion.__table__).values(rows)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_titulacion_generacion",
            set_={c: stmt.excluded[c] for c in [
                "matricula_generacional", "concluyeron_estudios", "egresados", "titulados",
            ]},
        )
        return session.execute(stmt).rowcount or len(rows)

    print("\nInsertando en base de datos...")
    with Session(engine) as session:
        try:
            n1 = upsert_matricula(session, matricula_rows)
            print(f"  ✓ Matrícula:       {n1} filas")
            n2 = insert_eval_academica(session, eval_acad_rows)
            print(f"  ✓ Eval. Académica: {n2} filas")
            n3 = insert_eval_docente(session, eval_doc_rows)
            print(f"  ✓ Eval. Docente:   {n3} filas")
            n4 = upsert_titulacion(session, tit_rows)
            print(f"  ✓ Titulación:      {n4} filas")
            session.commit()
            print("\n✅ Carga completada exitosamente.")
        except Exception as exc:
            session.rollback()
            print(f"\n❌ Error: {exc}")
            raise


def extract_all(wb: openpyxl.Workbook) -> tuple[list[dict], list[dict], list[dict], list[dict]]:
    print("Extrayendo datos...")
    matricula_rows = extract_matricula(wb)
    eval_acad_rows = extract_evaluacion_academica(wb)
    eval_doc_rows = extract_evaluacion_docente(wb)
    tit_rows = extract_titulacion(wb)

    print(f"  Matrícula:           {len(matricula_rows)} registros")
    print(f"  Eval. Académica:     {len(eval_acad_rows)} registros")
    print(f"  Eval. Docente:       {len(eval_doc_rows)} registros")
    print(f"  Titulación:          {len(tit_rows)} registros")
    _validation_summary(matricula_rows)
    return matricula_rows, eval_acad_rows, eval_doc_rows, tit_rows


def main() -> None:
    global _SUBSISTEMA_ID

    parser = argparse.ArgumentParser(description="Cargar indicadores históricos desde Excel")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Ruta al archivo Excel")
    parser.add_argument("--subsistema-id", type=int, default=5)
    parser.add_argument("--dry-run", action="store_true", help="Solo extraer y validar, sin escribir BD")
    parser.add_argument("--export-csv", type=Path, metavar="DIR", help="Exportar CSVs normalizados")
    args = parser.parse_args()

    _SUBSISTEMA_ID = args.subsistema_id

    excel_path = args.excel
    if not excel_path.exists():
        print(f"❌ No se encontró el archivo: {excel_path}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"Archivo: {excel_path}")
    print(f"Subsistema ID: {_SUBSISTEMA_ID}")
    if args.dry_run:
        print("Modo: DRY-RUN (sin escritura a BD)")
    print(f"{'='*60}\n")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    matricula_rows, eval_acad_rows, eval_doc_rows, tit_rows = extract_all(wb)
    wb.close()

    if args.export_csv:
        print(f"\nExportando CSVs a {args.export_csv}...")
        export_csv(args.export_csv, matricula_rows, eval_acad_rows, eval_doc_rows, tit_rows)

    if args.dry_run:
        print("\n✅ Dry-run completado.")
        return

    if args.export_csv and not args.dry_run:
        pass

    load_to_database(matricula_rows, eval_acad_rows, eval_doc_rows, tit_rows)


if __name__ == "__main__":
    main()