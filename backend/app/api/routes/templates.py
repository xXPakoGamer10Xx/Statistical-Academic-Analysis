"""Endpoint para descargar plantillas Excel de cada tipo de dataset."""
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, HTTPException
from fastapi.responses import Response
import openpyxl
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from app.schemas.upload import DatasetDefinitionOut, DatasetFieldOut
from app.services.dataset_definitions import DATASET_DEFINITIONS, get_dataset_definition

router = APIRouter()


@router.get("/formats", response_model=list[DatasetDefinitionOut])
async def list_dataset_formats() -> list[DatasetDefinitionOut]:
    return [
        DatasetDefinitionOut(
            key=definition.key,
            label=definition.label,
            description=definition.description,
            fields=[
                DatasetFieldOut(
                    name=field.name,
                    kind=field.kind,
                    required=field.required,
                    description=field.description,
                    allowed_values=list(field.allowed_values) if field.allowed_values else None,
                    suggested_values=list(field.suggested_values) if field.suggested_values else None,
                )
                for field in definition.fields
            ],
            catalogos={k: list(v) for k, v in definition.catalogos.items()}
            if definition.catalogos
            else None,
        )
        for definition in DATASET_DEFINITIONS.values()
    ]


@router.get("/{dataset_type}")
async def download_template(dataset_type: str) -> Response:
    """Genera y devuelve una plantilla Excel para el tipo de dataset solicitado."""
    try:
        definition = get_dataset_definition(dataset_type)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plantilla"

    # Filas de datos para las que se aplican las listas desplegables.
    LAST_DATA_ROW = 1000

    # Hoja "Listas" con los catalogos + rangos con nombre (DefinedName).
    # INDIRECT($Cn) resuelve el nombre definido cuyo nombre == valor de categoria en esa fila.
    # Funciona en Excel Desktop y Excel Online. Google Sheets no importa nombres definidos
    # desde .xlsx, por lo que en GS el desplegable de tipo quedara vacio; en ese caso
    # usar la captura manual de la app (que si tiene el desplegable dependiente nativo).
    listas_ws = wb.create_sheet("Listas")
    listas_col = 0

    # Mapa nombre_de_columna -> letra de columna en "Plantilla".
    letra_por_campo = {
        f.name: openpyxl.utils.get_column_letter(i)
        for i, f in enumerate(definition.fields, start=1)
    }

    # Columna "conductora": la que tiene los valores que son llaves del catalogo.
    driver_letter: str | None = None
    if definition.catalogos:
        llaves = set(definition.catalogos.keys())
        for field in definition.fields:
            if field.allowed_values and set(field.allowed_values) == llaves:
                driver_letter = letra_por_campo[field.name]
                break

        for categoria, valores in definition.catalogos.items():
            listas_col += 1
            cl = openpyxl.utils.get_column_letter(listas_col)
            listas_ws.cell(row=1, column=listas_col, value=categoria)
            for i, val in enumerate(valores, start=2):
                listas_ws.cell(row=i, column=listas_col, value=val)
            listas_ws.column_dimensions[cl].width = 20
            # Rango con nombre cuyo nombre == la categoria (ej. "beca" -> Listas!$A$2:$A$9).
            # INDIRECT($C2) en la validacion busca este nombre cuando C2 == "beca".
            ref = f"Listas!${cl}$2:${cl}${len(valores) + 1}"
            wb.defined_names[categoria] = DefinedName(categoria, attr_text=ref)

    # Escribir encabezados con estilo + validaciones de datos por columna
    for col_idx, field in enumerate(definition.fields, start=1):
        header = field.name
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = openpyxl.styles.PatternFill(
            start_color="1E3A5F", end_color="1E3A5F", fill_type="solid"
        )
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = max(len(header) + 4, 16)
        cell_range = f"{col_letter}2:{col_letter}{LAST_DATA_ROW}"

        if field.allowed_values:
            # Lista ESTRICTA: solo se permiten estos valores (ej. categoria: beca/discapacidad/etnia).
            opciones = ",".join(field.allowed_values)
            dv = DataValidation(
                type="list",
                formula1=f'"{opciones}"',
                allow_blank=True,
                showErrorMessage=True,
            )
            dv.errorTitle = "Valor no permitido"
            dv.error = f"Solo se permiten estas {len(field.allowed_values)} opciones: {opciones}."
            dv.promptTitle = header
            dv.prompt = f"Elige una opcion: {opciones}."
            ws.add_data_validation(dv)
            dv.add(cell_range)
        elif field.suggested_values and definition.catalogos and driver_letter:
            # Lista DEPENDIENTE via INDIRECT + rangos con nombre.
            # INDIRECT($C2) resuelve el rango con nombre igual al valor de categoria en esa fila.
            # Funciona en Excel Desktop y Excel Online. No bloquea (admite texto libre).
            dv = DataValidation(
                type="list",
                formula1=f"INDIRECT(${driver_letter}2)",
                allow_blank=True,
                showErrorMessage=False,
            )
            dv.promptTitle = header
            dv.prompt = "Primero elige la categoria; aqui aparecen solo sus tipos. Tambien puedes escribir el tuyo."
            ws.add_data_validation(dv)
            dv.add(cell_range)
        elif field.suggested_values:
            # Lista SUGERIDA plana: elegir del catalogo o escribir texto libre.
            listas_col += 1
            ref_letter = openpyxl.utils.get_column_letter(listas_col)
            listas_ws.cell(row=1, column=listas_col, value=header)
            for i, val in enumerate(field.suggested_values, start=2):
                listas_ws.cell(row=i, column=listas_col, value=val)
            last = len(field.suggested_values) + 1
            dv = DataValidation(
                type="list",
                formula1=f"Listas!${ref_letter}$2:${ref_letter}${last}",
                allow_blank=True,
                showErrorMessage=False,
            )
            dv.promptTitle = header
            dv.prompt = "Elige del catalogo sugerido o escribe tu propio valor."
            ws.add_data_validation(dv)
            dv.add(cell_range)

    # Guardar en memoria y devolver bytes completos para evitar ambiguedades en proxy/navegador
    buffer = BytesIO()
    wb.save(buffer)
    content = buffer.getvalue()

    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    filename = f"plantilla_{definition.key}_{stamp}.xlsx"
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"; filename*=UTF-8\'\'{filename}'
            ),
            "Content-Length": str(len(content)),
            "Access-Control-Expose-Headers": "Content-Disposition, Content-Length, Content-Type",
        },
    )
