"""Análisis inteligente de archivos Excel para detección automática de columnas."""
from __future__ import annotations

import difflib
import re
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl

from app.services.dataset_definitions import DATASET_DEFINITIONS

MAX_SHEETS = 10
MAX_COLS = 50
MAX_ROWS_SCAN = 100
SAMPLE_ROWS = 5
HEADER_SCAN_ROWS = 20
MATCH_THRESHOLD = 0.50
SUGGEST_THRESHOLD = 0.60


# ---------------------------------------------------------------------------
# Dataclasses de resultado
# ---------------------------------------------------------------------------

@dataclass
class ColumnMappingItem:
    excel_column: str
    system_field: str | None
    confidence: float


@dataclass
class DatasetTypeScore:
    dataset_type: str
    score: float
    label: str


@dataclass
class SheetAnalysis:
    sheet_name: str
    header_row: int
    detected_headers: list[str]
    sample_rows: list[list[str]]
    suggested_dataset_type: str | None
    dataset_type_scores: list[DatasetTypeScore]
    column_mapping: list[ColumnMappingItem]
    has_merged_cells: bool
    total_data_rows: int
    warnings: list[str]


@dataclass
class ExcelAnalysis:
    sheet_names: list[str]
    sheets: list[SheetAnalysis]
    recommended_sheet: str | None


# ---------------------------------------------------------------------------
# Normalización de nombres de columna
# ---------------------------------------------------------------------------

_DIACRITICS = str.maketrans({
    "á": "a", "à": "a", "ä": "a", "â": "a", "ã": "a", "å": "a",
    "é": "e", "è": "e", "ë": "e", "ê": "e",
    "í": "i", "ì": "i", "ï": "i", "î": "i",
    "ó": "o", "ò": "o", "ö": "o", "ô": "o", "õ": "o",
    "ú": "u", "ù": "u", "ü": "u", "û": "u",
    "ñ": "n", "ç": "c",
})


def _normalize(s: str) -> str:
    s = str(s).strip().lower()
    s = s.translate(_DIACRITICS)
    s = re.sub(r"[\s\-/\\]+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


# ---------------------------------------------------------------------------
# Fuzzy matching
# ---------------------------------------------------------------------------

def _match_score(excel_col: str, system_field: str) -> float:
    norm_excel = _normalize(excel_col)
    norm_field = system_field  # ya es snake_case limpio

    if norm_excel == norm_field:
        return 1.0
    if norm_excel in norm_field or norm_field in norm_excel:
        return 0.85
    return difflib.SequenceMatcher(None, norm_excel, norm_field).ratio()


def _best_match(excel_col: str, system_fields: list[str]) -> tuple[str | None, float]:
    best_field: str | None = None
    best_score = 0.0
    for sf in system_fields:
        score = _match_score(excel_col, sf)
        if score > best_score:
            best_score = score
            best_field = sf
    if best_score < MATCH_THRESHOLD:
        return None, best_score
    return best_field, best_score


# ---------------------------------------------------------------------------
# Detección de fila de headers
# ---------------------------------------------------------------------------

def _cell_value(cell: Any) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def _detect_header_row(ws: Any, max_scan: int = HEADER_SCAN_ROWS) -> int:
    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=1, max_row=max_scan):
        rows.append([_cell_value(c) for c in row])

    for i, row_vals in enumerate(rows[:-1]):
        non_empty = sum(1 for v in row_vals if v)
        if non_empty < 3:
            continue
        next_row = rows[i + 1]
        next_non_empty = sum(1 for v in next_row if v)
        if next_non_empty >= non_empty * 0.6:
            return i
    return 0


# ---------------------------------------------------------------------------
# Celdas combinadas
# ---------------------------------------------------------------------------

def _expand_merged_cells(ws: Any) -> bool:
    merged_ranges = list(ws.merged_cells.ranges)
    if not merged_ranges:
        return False
    for merged_range in merged_ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        top_left_value = ws.cell(min_row, min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                ws.cell(row, col).value = top_left_value
    return True


# ---------------------------------------------------------------------------
# Sugerencia de tipo de dataset
# ---------------------------------------------------------------------------

def _suggest_dataset_types(headers: list[str]) -> list[DatasetTypeScore]:
    scores: list[DatasetTypeScore] = []
    all_fields = [sf for definition in DATASET_DEFINITIONS.values() for sf in definition.required_columns]
    # para evitar calcular dos veces el mismo campo
    _ = all_fields

    for ds_type, definition in DATASET_DEFINITIONS.items():
        required = definition.required_columns
        if not required:
            continue
        matched = 0
        for sf in required:
            best_score = max((_match_score(h, sf) for h in headers), default=0.0)
            if best_score > SUGGEST_THRESHOLD:
                matched += 1
        score = matched / len(required)
        scores.append(DatasetTypeScore(
            dataset_type=ds_type,
            score=round(score, 3),
            label=definition.label,
        ))

    scores.sort(key=lambda x: x.score, reverse=True)
    return scores


# ---------------------------------------------------------------------------
# Análisis de una hoja
# ---------------------------------------------------------------------------

def _analyze_sheet(ws: Any, sheet_name: str) -> SheetAnalysis:
    warnings: list[str] = []

    has_merged = _expand_merged_cells(ws)
    if has_merged:
        warnings.append("La hoja contiene celdas combinadas — fueron normalizadas automáticamente.")

    header_row_idx = _detect_header_row(ws)
    if header_row_idx > 0:
        warnings.append(f"Los datos comienzan en la fila {header_row_idx + 1} (se omitieron filas de encabezado decorativas).")

    # Extraer headers (fila de headers)
    header_excel_row = header_row_idx + 1  # openpyxl es 1-indexed
    raw_headers: list[str] = []
    for cell in ws[header_excel_row]:
        val = _cell_value(cell)
        if val:
            raw_headers.append(val)
        if len(raw_headers) >= MAX_COLS:
            break

    if not raw_headers:
        return SheetAnalysis(
            sheet_name=sheet_name,
            header_row=header_row_idx,
            detected_headers=[],
            sample_rows=[],
            suggested_dataset_type=None,
            dataset_type_scores=[],
            column_mapping=[],
            has_merged_cells=has_merged,
            total_data_rows=0,
            warnings=warnings + ["No se detectaron columnas en esta hoja."],
        )

    n_cols = len(raw_headers)

    # Leer filas de datos
    data_rows: list[list[str]] = []
    row_num = header_excel_row + 1
    rows_scanned = 0
    for row in ws.iter_rows(min_row=row_num):
        if rows_scanned >= MAX_ROWS_SCAN:
            break
        vals = [_cell_value(row[i]) if i < len(row) else "" for i in range(n_cols)]
        non_empty = sum(1 for v in vals if v)
        if non_empty == 0:
            continue
        data_rows.append(vals)
        rows_scanned += 1

    sample_rows = [row for row in data_rows[:SAMPLE_ROWS]]
    total_data_rows = len(data_rows)

    # Sugerir tipo de dataset
    type_scores = _suggest_dataset_types(raw_headers)
    suggested = type_scores[0].dataset_type if type_scores and type_scores[0].score > 0 else None

    # Mapear columnas al mejor dataset sugerido (o al que tenga mayor score)
    if suggested and suggested in DATASET_DEFINITIONS:
        system_fields = [f.name for f in DATASET_DEFINITIONS[suggested].fields]
    else:
        system_fields = list({f.name for d in DATASET_DEFINITIONS.values() for f in d.fields})

    used_fields: set[str] = set()
    column_mapping: list[ColumnMappingItem] = []
    for excel_col in raw_headers:
        candidates = [(sf, _match_score(excel_col, sf)) for sf in system_fields if sf not in used_fields]
        candidates.sort(key=lambda x: x[1], reverse=True)
        if candidates and candidates[0][1] >= MATCH_THRESHOLD:
            best_field, best_score = candidates[0]
            used_fields.add(best_field)
            column_mapping.append(ColumnMappingItem(
                excel_column=excel_col,
                system_field=best_field,
                confidence=round(best_score, 3),
            ))
        else:
            column_mapping.append(ColumnMappingItem(
                excel_column=excel_col,
                system_field=None,
                confidence=0.0,
            ))

    return SheetAnalysis(
        sheet_name=sheet_name,
        header_row=header_row_idx,
        detected_headers=raw_headers,
        sample_rows=sample_rows,
        suggested_dataset_type=suggested,
        dataset_type_scores=type_scores,
        column_mapping=column_mapping,
        has_merged_cells=has_merged,
        total_data_rows=total_data_rows,
        warnings=warnings,
    )


# ---------------------------------------------------------------------------
# Función principal
# ---------------------------------------------------------------------------

def analyze_excel_file(file_path: str) -> ExcelAnalysis:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames

    sheets_to_analyze = sheet_names[:MAX_SHEETS]
    sheet_analyses: list[SheetAnalysis] = []

    for name in sheets_to_analyze:
        ws = wb[name]
        analysis = _analyze_sheet(ws, name)
        sheet_analyses.append(analysis)

    wb.close()

    # La hoja recomendada es la que tiene mayor score de dataset type
    recommended: str | None = None
    best_score = 0.0
    for sa in sheet_analyses:
        if sa.dataset_type_scores and sa.dataset_type_scores[0].score > best_score and sa.total_data_rows > 0:
            best_score = sa.dataset_type_scores[0].score
            recommended = sa.sheet_name

    return ExcelAnalysis(
        sheet_names=sheet_names,
        sheets=sheet_analyses,
        recommended_sheet=recommended,
    )


def analyze_excel_bytes(file_bytes: bytes, suffix: str = ".xlsx") -> ExcelAnalysis:
    """Analiza un archivo Excel a partir de bytes (para uso en endpoints HTTP)."""
    tmp = Path(tempfile.mktemp(suffix=suffix))
    try:
        tmp.write_bytes(file_bytes)
        return analyze_excel_file(str(tmp))
    finally:
        tmp.unlink(missing_ok=True)
