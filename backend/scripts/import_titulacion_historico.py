"""ETL: carga Titulacion historica de UPTex (subsistema_id=5) desde la hoja
"Titulados Histórico act" del Excel institucional, hacia la tabla `titulacion`
(con desglose H/M agregado en 0017_titulacion_sexo_breakdown).

Cada fila de la hoja es una micro-cohorte (una fecha exacta de ingreso/egreso). Se
agrupan por carrera + (año de ingreso - año de egreso) en una "generacion", sumando
ingreso/egresados/titulados de todas las micro-cohortes de esa generacion — mismo
criterio que usan las hojas resumen "Eficiencia Terminal Historico"/"Tasa de
Titulación" (una columna "M <año>" por generacion).

matricula_generacional = suma de INGRESO (H+M) de la generacion.
egresados = suma de EGRESADOS TOTAL (con su cohorte + rezagados, H+M).
titulados = suma de TITULADOS TOTAL (H+M).
concluyeron_estudios se deja NULL (no viene en esta hoja; ya no se captura por
formulario/CSV tampoco, ver comentario en el modelo Titulacion).

Por defecto corre en modo DRY-RUN. Usar --apply para escribir de verdad.
"""
from __future__ import annotations

import argparse
import asyncio
from pathlib import Path

import openpyxl
from sqlalchemy import insert

from app.core.database import AsyncSessionLocal
from app.models.titulacion import Titulacion
from scripts.uptex_carreras import normalize_carrera_fullname

SUBSISTEMA_ID = 5
SHEET_NAME = "Titulados Histórico act"
FIRST_ROW = 18
LAST_ROW = 61

# columnas (1-indexadas) segun el encabezado de 3 filas (15/16/17) de la hoja
COL_CARRERA = 2
COL_INGRESO_FECHA = 3
COL_EGRESO_FECHA = 4
COL_H_INGRESO = 5
COL_M_INGRESO = 6
COL_H_EGR_COHORTE = 8
COL_M_EGR_COHORTE = 9
COL_H_EGR_REZAGADOS = 10
COL_M_EGR_REZAGADOS = 11
COL_H_TITULADOS = 13
COL_M_TITULADOS = 14


def _n(v) -> int:
    return int(v) if v is not None else 0


def parse_titulados_historico(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]

    acumulado: dict[tuple[str, str], dict[str, int]] = {}
    carrera_actual: str | None = None
    skipped_unknown: set[str] = set()
    skipped_sin_fecha = 0

    for r in range(FIRST_ROW, LAST_ROW + 1):
        raw_carrera = ws.cell(row=r, column=COL_CARRERA).value
        if isinstance(raw_carrera, str) and raw_carrera.strip():
            if raw_carrera.strip().upper() == "TOTAL":
                continue
            carrera_actual = raw_carrera.strip()

        ingreso_fecha = ws.cell(row=r, column=COL_INGRESO_FECHA).value
        egreso_fecha = ws.cell(row=r, column=COL_EGRESO_FECHA).value
        if ingreso_fecha is None or egreso_fecha is None or carrera_actual is None:
            continue

        programa = normalize_carrera_fullname(carrera_actual)
        if programa is None:
            skipped_unknown.add(carrera_actual)
            continue

        generacion = f"{ingreso_fecha.year}-{egreso_fecha.year}"
        key = (programa, generacion)
        acc = acumulado.setdefault(key, {
            "ingreso_h": 0, "ingreso_m": 0,
            "egresados_h": 0, "egresados_m": 0,
            "titulados_h": 0, "titulados_m": 0,
        })
        acc["ingreso_h"] += _n(ws.cell(row=r, column=COL_H_INGRESO).value)
        acc["ingreso_m"] += _n(ws.cell(row=r, column=COL_M_INGRESO).value)
        acc["egresados_h"] += _n(ws.cell(row=r, column=COL_H_EGR_COHORTE).value) + _n(
            ws.cell(row=r, column=COL_H_EGR_REZAGADOS).value)
        acc["egresados_m"] += _n(ws.cell(row=r, column=COL_M_EGR_COHORTE).value) + _n(
            ws.cell(row=r, column=COL_M_EGR_REZAGADOS).value)
        acc["titulados_h"] += _n(ws.cell(row=r, column=COL_H_TITULADOS).value)
        acc["titulados_m"] += _n(ws.cell(row=r, column=COL_M_TITULADOS).value)

    if skipped_unknown:
        print(f"[aviso] carreras no reconocidas, omitidas: {sorted(skipped_unknown)}")
    if skipped_sin_fecha:
        print(f"[aviso] {skipped_sin_fecha} fila(s) sin fecha de ingreso/egreso omitidas.")

    rows = []
    for (programa, generacion), acc in acumulado.items():
        rows.append({
            "subsistema_id": SUBSISTEMA_ID,
            "generacion": generacion,
            "programa_educativo": programa,
            "matricula_generacional": acc["ingreso_h"] + acc["ingreso_m"],
            "concluyeron_estudios": None,
            "egresados": acc["egresados_h"] + acc["egresados_m"],
            "titulados": acc["titulados_h"] + acc["titulados_m"],
            "ingresados_ns": None,
            "egresados_hombres": acc["egresados_h"],
            "egresados_mujeres": acc["egresados_m"],
            "titulados_hombres": acc["titulados_h"],
            "titulados_mujeres": acc["titulados_m"],
        })
    return rows


async def load_rows(rows: list[dict], apply: bool) -> None:
    rows_sorted = sorted(rows, key=lambda r: (r["programa_educativo"], r["generacion"]))
    print(f"\n{'GENERACION':<12} {'PE':<55} {'INGRESO':>8} {'EGRESADOS':>10} {'TITULADOS':>10}")
    for r in rows_sorted:
        print(f"{r['generacion']:<12} {r['programa_educativo']:<55} "
              f"{r['matricula_generacional']:>8} {r['egresados']:>10} {r['titulados']:>10}")
    print(f"\nTotal filas a cargar: {len(rows)}")

    if not apply:
        print("\n[DRY RUN] No se escribió nada. Corre con --apply para cargar de verdad.")
        return

    async with AsyncSessionLocal() as db:
        from sqlalchemy import func, select

        existing = (await db.execute(
            select(func.count()).select_from(Titulacion).where(Titulacion.subsistema_id == SUBSISTEMA_ID)
        )).scalar_one()
        if existing:
            print(f"\n[ABORTADO] Ya existen {existing} filas de titulacion para "
                  f"subsistema_id={SUBSISTEMA_ID}. Bórralas primero si quieres recargar.")
            return

        result = await db.execute(insert(Titulacion.__table__), rows)
        await db.commit()
        print(f"\n[APLICADO] filas insertadas en titulacion (rowcount driver: {result.rowcount}).")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    rows = parse_titulados_historico(args.excel)
    asyncio.run(load_rows(rows, args.apply))


if __name__ == "__main__":
    main()
