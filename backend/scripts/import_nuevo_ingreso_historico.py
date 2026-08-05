"""ETL: carga Nuevo Ingreso historico de UPTex (subsistema_id=5) desde la hoja
"Examen Admision" del Excel institucional, hacia la tabla `matricula`.

Por defecto corre en modo DRY-RUN (solo imprime lo que cargaria). Usar --apply
para escribir de verdad.

Uso (dentro del contenedor backend):
    python -m scripts.import_nuevo_ingreso_historico --excel /data/uploads/historico.xlsx
    python -m scripts.import_nuevo_ingreso_historico --excel /data/uploads/historico.xlsx --apply
"""
from __future__ import annotations

import argparse
import asyncio
import sys
from pathlib import Path

import openpyxl

from app.core.database import AsyncSessionLocal
from app.models.matricula import Matricula
from app.services.upload_validations import (
    apply_matricula_carry_forward,
    build_previous_matricula_stmt,
    matricula_carry_forward_lookups,
)
from scripts.uptex_carreras import carrera_por_codigo

SUBSISTEMA_ID = 5
SHEET_NAME = "Examen Admision"
HEADER_ROW = 10  # fila con 'PE','Examen','RENOES'/'Hagamoslo Juntos','UAEM-GEM','Pase Directo','TOTAL'
CICLO_LABEL_ROW = 8  # fila con 'NUEVO INGRESO 2019-2020', etc.


def _find_blocks(ws) -> list[dict]:
    """Encuentra cada bloque de ciclo: columna donde empieza 'PE' y su etiqueta de ciclo."""
    header_cells = list(ws[HEADER_ROW])
    ciclo_cells = list(ws[CICLO_LABEL_ROW])

    ciclo_by_col: dict[int, str] = {}
    for cell in ciclo_cells:
        if isinstance(cell.value, str) and "NUEVO INGRESO" in cell.value.upper():
            ciclo = cell.value.upper().replace("NUEVO INGRESO", "").strip()
            ciclo_by_col[cell.column] = ciclo

    blocks = []
    pe_cols = [c.column for c in header_cells if isinstance(c.value, str) and c.value.strip().upper() == "PE"]
    for pe_col in pe_cols:
        # etiqueta de ciclo: la mas cercana en o despues de esta columna (normalmente pe_col+1)
        candidatos = [col for col in ciclo_by_col if col >= pe_col]
        if not candidatos:
            continue
        ciclo_col = min(candidatos)
        col_names = [ws.cell(row=HEADER_ROW, column=pe_col + i).value for i in range(6)]
        col_names = [str(c).strip().upper().replace("\n", " ") if c else "" for c in col_names]
        blocks.append({
            "pe_col": pe_col,
            "ciclo_escolar": ciclo_by_col[ciclo_col],
            "col_names": col_names,  # [PE, Examen, RENOES/Hagamoslo, UAEM-GEM, Pase Directo, TOTAL]
        })
    return blocks


def parse_examen_admision(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    blocks = _find_blocks(ws)

    rows: list[dict] = []
    skipped_lagp = 0
    skipped_unknown: set[str] = set()

    for block in blocks:
        pe_col = block["pe_col"]
        ciclo_escolar = block["ciclo_escolar"]
        for r in range(HEADER_ROW + 1, ws.max_row + 1):
            codigo = ws.cell(row=r, column=pe_col).value
            if not isinstance(codigo, str) or not codigo.strip():
                continue
            codigo = codigo.strip().upper()
            if codigo in ("ACEPTADOS", "TOTAL"):
                continue

            examen = ws.cell(row=r, column=pe_col + 1).value
            col3 = ws.cell(row=r, column=pe_col + 2).value  # RENOES o Hagamoslo Juntos
            uaem_gem = ws.cell(row=r, column=pe_col + 3).value
            pase_directo = ws.cell(row=r, column=pe_col + 4).value

            if codigo == "LAGP":
                if any(v not in (None, 0) for v in (examen, col3, uaem_gem, pase_directo)):
                    skipped_lagp += 1
                continue

            programa = carrera_por_codigo(codigo)
            if programa is None:
                if any(v not in (None, 0) for v in (examen, col3, uaem_gem, pase_directo)):
                    skipped_unknown.add(codigo)
                continue

            rows.append({
                "subsistema_id": SUBSISTEMA_ID,
                "ciclo_escolar": ciclo_escolar,
                "cuatrimestre": 1,  # el ingreso siempre entra en el cuatrimestre Sep-Dic
                "programa_educativo": programa,
                "ingreso_examen": int(examen or 0),
                "ingreso_renoes": int(col3 or 0),  # incluye "Hagamoslo Juntos" (nombre anterior de RENOES)
                "ingreso_uaem_gem": int(uaem_gem or 0),
                "ingreso_pase_directo": int(pase_directo or 0),
            })

    if skipped_lagp:
        print(f"[aviso] {skipped_lagp} fila(s) de 'LAGP' con datos no cero se omitieron "
              f"(nombre completo sin confirmar).", file=sys.stderr)
    if skipped_unknown:
        print(f"[aviso] codigos no reconocidos con datos: {sorted(skipped_unknown)}", file=sys.stderr)

    # Si dos filas del mismo bloque terminan mapeando al mismo nombre final (no deberia
    # pasar en este sheet, pero por seguridad), se suman en vez de sobreescribirse.
    merged: dict[tuple, dict] = {}
    for row in rows:
        key = (row["ciclo_escolar"], row["cuatrimestre"], row["programa_educativo"])
        if key in merged:
            for f in ("ingreso_examen", "ingreso_renoes", "ingreso_uaem_gem", "ingreso_pase_directo"):
                merged[key][f] += row[f]
        else:
            merged[key] = dict(row)
    return list(merged.values())


async def load_rows(rows: list[dict], apply: bool) -> None:
    rows_sorted = sorted(rows, key=lambda r: (r["programa_educativo"], r["ciclo_escolar"], r["cuatrimestre"]))

    print(f"\n{'CICLO':<12} {'PE':<55} {'EXAMEN':>7} {'RENOES':>7} {'UAEM-GEM':>9} {'PASE DIR':>9}")
    for r in rows_sorted:
        print(f"{r['ciclo_escolar']:<12} {r['programa_educativo']:<55} "
              f"{r['ingreso_examen']:>7} {r['ingreso_renoes']:>7} {r['ingreso_uaem_gem']:>9} {r['ingreso_pase_directo']:>9}")
    print(f"\nTotal filas a cargar: {len(rows)}")

    if not apply:
        print("\n[DRY RUN] No se escribió nada. Corre con --apply para cargar de verdad.")
        return

    async with AsyncSessionLocal() as db:
        lookups = matricula_carry_forward_lookups(rows)
        previous_by_key: dict[tuple, dict[str, int] | None] = {}
        for (subsistema_id, programa_educativo), (ciclo_escolar, cuatrimestre) in lookups.items():
            stmt = build_previous_matricula_stmt(subsistema_id, programa_educativo, ciclo_escolar, cuatrimestre)
            previous_row = (await db.execute(stmt)).first()
            previous_by_key[(subsistema_id, programa_educativo)] = (
                {
                    "total": previous_row.total,
                    "bajas_reprobacion": previous_row.bajas_reprobacion,
                    "bajas_desercion": previous_row.bajas_desercion,
                    "nuevo_ingreso": previous_row.nuevo_ingreso,
                }
                if previous_row is not None
                else None
            )
        apply_matricula_carry_forward(rows, previous_by_key)

        from sqlalchemy.dialects.postgresql import insert as pg_insert

        stmt = pg_insert(Matricula.__table__).values(rows)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_matricula_periodo",
            set_={c: stmt.excluded[c] for c in [
                "total", "nuevo_ingreso", "ingreso_examen", "ingreso_pase_directo",
                "ingreso_renoes", "ingreso_uaem_gem",
            ]},
        )
        result = await db.execute(stmt)
        await db.commit()
        print(f"\n[APLICADO] {result.rowcount} filas insertadas/actualizadas en matricula.")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    rows = parse_examen_admision(args.excel)
    asyncio.run(load_rows(rows, args.apply))


if __name__ == "__main__":
    main()
