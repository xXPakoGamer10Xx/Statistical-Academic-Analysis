"""ETL: carga Aprovechamiento historico de UPTex (subsistema_id=5) desde la hoja
"APROVECHAMIENTO" del Excel institucional, hacia la tabla `evaluacion_academica`.

num_pe se deja en 1 para todas las filas (peso uniforme): el Excel no trae el numero
de materias/PE evaluados por cuatrimestre, solo el promedio ya calculado.

Por defecto corre en modo DRY-RUN. Usar --apply para escribir de verdad.
"""
from __future__ import annotations

import argparse
import asyncio
from pathlib import Path

import openpyxl
from sqlalchemy import insert

from app.core.database import AsyncSessionLocal
from app.models.evaluacion import EvaluacionAcademica
from scripts.uptex_carreras import carrera_por_codigo

SUBSISTEMA_ID = 5
SHEET_NAME = "APROVECHAMIENTO"
YEAR_ROW = 5
SUBHEADER_ROW = 6
FIRST_PE_ROW = 7
LAST_PE_ROW = 13


def _ciclo_cuatrimestre(periodo: str, year: int) -> tuple[str, int]:
    periodo = periodo.strip().upper()
    if periodo.startswith("S"):  # Sep-Dic
        return f"{year}-{year + 1}", 1
    if periodo.startswith("E"):  # Ene-Abr
        return f"{year - 1}-{year}", 2
    if periodo.startswith("M"):  # May-Ago
        return f"{year - 1}-{year}", 3
    raise ValueError(f"periodo no reconocido: {periodo!r}")


def parse_aprovechamiento(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]

    # columna -> año (solo en la 1ra columna de cada bloque de 3)
    year_by_col: dict[int, int] = {
        c.column: c.value for c in ws[YEAR_ROW] if isinstance(c.value, int)
    }
    # para cada columna de dato, su año de bloque (la col de año <= col actual, la mas cercana)
    year_cols_sorted = sorted(year_by_col)

    def year_for_col(col: int) -> int | None:
        candidatos = [c for c in year_cols_sorted if c <= col]
        return year_by_col[max(candidatos)] if candidatos else None

    subheaders = {c.column: c.value for c in ws[SUBHEADER_ROW] if isinstance(c.value, str)}

    rows: list[dict] = []
    skipped_lagp = 0
    skipped_unknown: set[str] = set()

    for r in range(FIRST_PE_ROW, LAST_PE_ROW + 1):
        codigo = ws.cell(row=r, column=2).value
        if not isinstance(codigo, str) or not codigo.strip():
            continue
        codigo = codigo.strip().upper()
        programa = carrera_por_codigo(codigo)

        for col, label in subheaders.items():
            label = label.strip().upper().replace("\n", "").replace(" ", "")
            if label.startswith("PROM"):
                continue  # columna resumen, no es un cuatrimestre real
            year = year_for_col(col)
            if year is None:
                continue
            promedio = ws.cell(row=r, column=col).value
            if promedio is None:
                continue

            if codigo == "LAGP":
                skipped_lagp += 1
                continue
            if programa is None:
                skipped_unknown.add(codigo)
                continue

            ciclo_escolar, cuatrimestre = _ciclo_cuatrimestre(label, year)
            rows.append({
                "subsistema_id": SUBSISTEMA_ID,
                "ciclo_escolar": ciclo_escolar,
                "cuatrimestre": cuatrimestre,
                "programa_educativo": programa,
                "promedio_pe": round(float(promedio), 2),
                "num_pe": 1,
            })

    if skipped_lagp:
        print(f"[aviso] {skipped_lagp} valor(es) de 'LAGP' omitidos (nombre completo sin confirmar).")
    if skipped_unknown:
        print(f"[aviso] codigos no reconocidos con datos: {sorted(skipped_unknown)}")

    # Si dos codigos distintos (viejo/nuevo plan) cayeran en el mismo periodo para la
    # misma carrera final (no pasa en esta hoja, pero por seguridad), promediar en vez
    # de sobreescribir silenciosamente.
    merged: dict[tuple, list[dict]] = {}
    for row in rows:
        key = (row["ciclo_escolar"], row["cuatrimestre"], row["programa_educativo"])
        merged.setdefault(key, []).append(row)
    out = []
    for key, group in merged.items():
        if len(group) == 1:
            out.append(group[0])
        else:
            avg = sum(g["promedio_pe"] for g in group) / len(group)
            merged_row = dict(group[0])
            merged_row["promedio_pe"] = round(avg, 2)
            merged_row["num_pe"] = len(group)
            out.append(merged_row)
    return out


async def load_rows(rows: list[dict], apply: bool) -> None:
    rows_sorted = sorted(rows, key=lambda r: (r["programa_educativo"], r["ciclo_escolar"], r["cuatrimestre"]))
    print(f"\n{'CICLO':<12} {'CUATRI':<7} {'PE':<55} {'PROMEDIO':>9}")
    for r in rows_sorted:
        print(f"{r['ciclo_escolar']:<12} {r['cuatrimestre']:<7} {r['programa_educativo']:<55} {r['promedio_pe']:>9}")
    print(f"\nTotal filas a cargar: {len(rows)}")

    if not apply:
        print("\n[DRY RUN] No se escribió nada. Corre con --apply para cargar de verdad.")
        return

    async with AsyncSessionLocal() as db:
        # No hay unique constraint en esta tabla (la app tampoco hace upsert para este
        # dataset): se hace INSERT simple, protegido con un chequeo de idempotencia para
        # no duplicar si el script se corre dos veces por error.
        from sqlalchemy import func, select

        existing = (await db.execute(
            select(func.count()).select_from(EvaluacionAcademica).where(
                EvaluacionAcademica.subsistema_id == SUBSISTEMA_ID
            )
        )).scalar_one()
        if existing:
            print(f"\n[ABORTADO] Ya existen {existing} filas de evaluacion_academica para "
                  f"subsistema_id={SUBSISTEMA_ID}. Bórralas primero si quieres recargar "
                  f"(este dataset no tiene upsert), para no duplicar.")
            return

        result = await db.execute(insert(EvaluacionAcademica.__table__), rows)
        await db.commit()
        print(f"\n[APLICADO] {result.rowcount} filas insertadas en evaluacion_academica.")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    rows = parse_aprovechamiento(args.excel)
    asyncio.run(load_rows(rows, args.apply))


if __name__ == "__main__":
    main()
