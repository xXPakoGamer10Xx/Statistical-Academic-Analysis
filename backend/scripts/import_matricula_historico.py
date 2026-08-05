"""ETL: completa Matricula historica de UPTex (subsistema_id=5) con bajas por
reprobacion/desercion y hombres/mujeres POR CARRERA, cuatrimestre por cuatrimestre
(2019-2020 a 2024-2025).

El Excel institucional NO trae bajas ni hombres/mujeres desglosados por carrera —
solo totales institucionales por periodo (hojas "REP Y DES" y "Matricula por
Género"). Por eso estos valores se ESTIMAN repartiendo el total institucional de
cada periodo proporcionalmente a la matricula de cada carrera en ese mismo periodo
(hoja "Matricula por cuatrimestre"), y se guardan como estimacion, no como dato
reportado por carrera (confirmado con el usuario).

Como esta carga agrega registros para TODOS los cuatrimestres (no solo el 1, que ya
tenia datos de nuevo ingreso), se recalcula el arrastre completo de Matricula Actual
(total) cuatrimestre a cuatrimestre para cada carrera, en orden cronologico.

Por defecto corre en modo DRY-RUN. Usar --apply para escribir de verdad.
"""
from __future__ import annotations

import argparse
import asyncio
from pathlib import Path

import openpyxl
from sqlalchemy import select
from sqlalchemy.dialects.postgresql import insert as pg_insert

from app.core.database import AsyncSessionLocal
from app.models.matricula import Matricula
from app.services.formulas import calculate_matricula_actual
from scripts.uptex_carreras import carrera_por_codigo

SUBSISTEMA_ID = 5
MATRICULA_SHEET = "Matricula por cuatrimestre"
REPDES_SHEET = "REP Y DES"
GENERO_SHEET = "Matricula por Género"


def _periodo_a_ciclo(label: str) -> tuple[str, int] | None:
    """'Sep-Dic 2019' / 'S-D 2019' / 'E-A 2020' / 'M - A 2020' -> (ciclo_escolar, cuatrimestre)."""
    s = label.strip().upper().replace(" ", "").replace("-", "")
    import re
    m = re.search(r"(\d{4})", s)
    if not m:
        return None
    year = int(m.group(1))
    if s.startswith("SEPDIC") or s.startswith("SD"):
        return f"{year}-{year + 1}", 1
    if s.startswith("ENEABR") or s.startswith("EA"):
        return f"{year - 1}-{year}", 2
    if s.startswith("MAYAGO") or s.startswith("MA"):
        return f"{year - 1}-{year}", 3
    return None


def parse_matricula_por_carrera(xlsx_path: Path, wb) -> dict[tuple[str, int, str], int]:
    """(ciclo, cuatrimestre, programa) -> matricula total de esa carrera en ese periodo."""
    ws = wb[MATRICULA_SHEET]
    period_by_col: dict[int, tuple[str, int]] = {}
    for c in ws[6]:
        if isinstance(c.value, str):
            parsed = _periodo_a_ciclo(c.value)
            if parsed:
                period_by_col[c.column] = parsed

    out: dict[tuple[str, int, str], int] = {}
    for r in range(7, 20):
        codigo = ws.cell(row=r, column=2).value
        if not isinstance(codigo, str) or not codigo.strip():
            continue
        programa = carrera_por_codigo(codigo.strip().upper())
        if programa is None:
            continue
        for col, (ciclo, cuatri) in period_by_col.items():
            val = ws.cell(row=r, column=col).value
            if not val:
                continue
            key = (ciclo, cuatri, programa)
            out[key] = out.get(key, 0) + int(val)
    return out


def parse_institucional_bajas(wb) -> dict[tuple[str, int], tuple[int, int]]:
    """(ciclo, cuatrimestre) -> (bajas_desercion, bajas_reprobacion) institucional total."""
    ws = wb[REPDES_SHEET]
    labels = list(ws[6])
    desercion = list(ws[8])
    reprobacion = list(ws[9])
    out: dict[tuple[str, int], tuple[int, int]] = {}
    for label_cell, des_cell, rep_cell in zip(labels, desercion, reprobacion):
        if not isinstance(label_cell.value, str):
            continue
        parsed = _periodo_a_ciclo(label_cell.value)
        if not parsed:
            continue
        if des_cell.value is None or rep_cell.value is None:
            continue
        out[parsed] = (int(des_cell.value), int(rep_cell.value))
    return out


def parse_institucional_genero(wb) -> dict[tuple[str, int], tuple[int, int]]:
    """(ciclo, cuatrimestre) -> (hombres, mujeres) institucional total."""
    ws = wb[GENERO_SHEET]
    labels = list(ws[10])
    mujeres = list(ws[11])
    hombres = list(ws[12])
    out: dict[tuple[str, int], tuple[int, int]] = {}
    for label_cell, m_cell, h_cell in zip(labels, mujeres, hombres):
        if not isinstance(label_cell.value, str):
            continue
        parsed = _periodo_a_ciclo(label_cell.value)
        if not parsed:
            continue
        if m_cell.value is None or h_cell.value is None:
            continue
        out[parsed] = (int(h_cell.value), int(m_cell.value))
    return out


def build_rows(xlsx_path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    matricula_por_carrera = parse_matricula_por_carrera(xlsx_path, wb)
    bajas_inst = parse_institucional_bajas(wb)
    genero_inst = parse_institucional_genero(wb)

    # total institucional por periodo (suma de todas las carreras) = denominador del reparto
    total_por_periodo: dict[tuple[str, int], int] = {}
    for (ciclo, cuatri, _programa), val in matricula_por_carrera.items():
        key = (ciclo, cuatri)
        total_por_periodo[key] = total_por_periodo.get(key, 0) + val

    rows_by_key: dict[tuple[str, int, str], dict] = {}
    periodos_sin_bajas: set[tuple[str, int]] = set()
    periodos_sin_genero: set[tuple[str, int]] = set()

    for (ciclo, cuatri, programa), carrera_total in matricula_por_carrera.items():
        periodo = (ciclo, cuatri)
        total_periodo = total_por_periodo[periodo]
        peso = carrera_total / total_periodo if total_periodo else 0

        bajas_reprobacion = bajas_desercion = 0
        if periodo in bajas_inst:
            desercion_inst, reprobacion_inst = bajas_inst[periodo]
            bajas_desercion = round(desercion_inst * peso)
            bajas_reprobacion = round(reprobacion_inst * peso)
        else:
            periodos_sin_bajas.add(periodo)

        hombres = mujeres = 0
        if periodo in genero_inst:
            hombres_inst, mujeres_inst = genero_inst[periodo]
            hombres = round(hombres_inst * peso)
            mujeres = round(mujeres_inst * peso)
        else:
            periodos_sin_genero.add(periodo)

        rows_by_key[(ciclo, cuatri, programa)] = {
            "subsistema_id": SUBSISTEMA_ID,
            "ciclo_escolar": ciclo,
            "cuatrimestre": cuatri,
            "programa_educativo": programa,
            "bajas_reprobacion": bajas_reprobacion,
            "bajas_desercion": bajas_desercion,
            "hombres": hombres,
            "mujeres": mujeres,
        }

    if periodos_sin_bajas:
        print(f"[aviso] sin dato institucional de bajas para {len(periodos_sin_bajas)} periodo(s) "
              f"(quedan en 0): {sorted(periodos_sin_bajas)}")
    if periodos_sin_genero:
        print(f"[aviso] sin dato institucional de género para {len(periodos_sin_genero)} periodo(s) "
              f"(quedan en 0): {sorted(periodos_sin_genero)}")

    return list(rows_by_key.values())


async def load_rows(rows: list[dict], apply: bool) -> None:
    rows_sorted = sorted(rows, key=lambda r: (r["programa_educativo"], r["ciclo_escolar"], r["cuatrimestre"]))
    print(f"\n{'CICLO':<12} {'CUATRI':<7} {'PE':<50} {'BAJ.REP':>8} {'BAJ.DES':>8} {'H':>6} {'M':>6}")
    for r in rows_sorted:
        print(f"{r['ciclo_escolar']:<12} {r['cuatrimestre']:<7} {r['programa_educativo']:<50} "
              f"{r['bajas_reprobacion']:>8} {r['bajas_desercion']:>8} {r['hombres']:>6} {r['mujeres']:>6}")
    print(f"\nTotal filas: {len(rows)}")

    if not apply:
        print("\n[DRY RUN] No se escribió nada. Corre con --apply para cargar de verdad.")
        return

    async with AsyncSessionLocal() as db:
        # Traer nuevo_ingreso ya cargado (solo existe en cuatrimestre=1, por Nuevo Ingreso ETL)
        existing = (await db.execute(
            select(Matricula.ciclo_escolar, Matricula.cuatrimestre, Matricula.programa_educativo, Matricula.nuevo_ingreso)
            .where(Matricula.subsistema_id == SUBSISTEMA_ID)
        )).all()
        nuevo_ingreso_by_key = {(r.ciclo_escolar, r.cuatrimestre, r.programa_educativo): r.nuevo_ingreso for r in existing}

        by_row = {(r["ciclo_escolar"], r["cuatrimestre"], r["programa_educativo"]): r for r in rows}

        # arrastre completo, cronologico, cuatrimestre a cuatrimestre, por carrera
        by_programa: dict[str, list[tuple[str, int]]] = {}
        for ciclo, cuatri, programa in by_row:
            by_programa.setdefault(programa, []).append((ciclo, cuatri))

        for programa, periodos in by_programa.items():
            periodos.sort(key=lambda p: (p[0], p[1]))
            last_total = 0
            for ciclo, cuatri in periodos:
                key = (ciclo, cuatri, programa)
                row = by_row[key]
                nuevo_ingreso = nuevo_ingreso_by_key.get(key, 0)
                row["total"] = last_total
                last_total = calculate_matricula_actual(
                    last_total, row["bajas_reprobacion"], row["bajas_desercion"], nuevo_ingreso
                )

        stmt = pg_insert(Matricula.__table__).values(list(by_row.values()))
        stmt = stmt.on_conflict_do_update(
            constraint="uq_matricula_periodo",
            set_={c: stmt.excluded[c] for c in [
                "total", "bajas_reprobacion", "bajas_desercion", "hombres", "mujeres",
            ]},
        )
        result = await db.execute(stmt)
        await db.commit()
        print(f"\n[APLICADO] filas insertadas/actualizadas en matricula (rowcount driver: {result.rowcount}).")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--excel", required=True, type=Path)
    parser.add_argument("--apply", action="store_true")
    args = parser.parse_args()

    rows = build_rows(args.excel)
    asyncio.run(load_rows(rows, args.apply))


if __name__ == "__main__":
    main()
